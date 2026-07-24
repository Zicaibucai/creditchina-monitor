import tempfile
import unittest
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from creditchina_merged.crawler import EnterpriseRecord
from creditchina_merged.exporter import XlsxExporter


class XlsxExporterTests(unittest.TestCase):
    def test_exports_current_company_to_separate_section_sheets(self):
        record = EnterpriseRecord(
            name="示例有限公司",
            encry_str="key",
            basic={"法人": "张三", "统一社会信用代码": "CODE"},
            permissions=[{"行政许可决定书文号": "许可1"}],
            penalties=[{"决定书文号": "处罚1", "处罚结果": "罚款"}],
            red_list=[],
            watch_list=[],
            black_list=[],
        )
        with tempfile.TemporaryDirectory() as directory:
            path = XlsxExporter.export_current_company(
                record,
                Path(directory) / "current.xlsx",
            )
            workbook = load_workbook(path, read_only=False)

        self.assertEqual(
            workbook.sheetnames[:4],
            ["导出说明", "基本信息", "行政许可", "行政处罚"],
        )
        self.assertEqual(workbook["基本信息"]["A5"].value, "示例有限公司")
        self.assertEqual(workbook["行政处罚"].freeze_panes, "A5")

    def test_history_penalty_export_keeps_old_and_new_records(self):
        observed = datetime(2026, 7, 16, 12, 0, 0)
        rows = [
            {
                "company_name": "示例有限公司",
                "section_name": "行政处罚",
                "record_key": "旧处罚",
                "raw_json": '{"决定书文号": "旧处罚"}',
                "first_seen_at": observed,
                "last_seen_at": observed,
                "seen_count": 1,
            },
            {
                "company_name": "示例有限公司",
                "section_name": "行政处罚",
                "record_key": "新处罚",
                "raw_json": '{"决定书文号": "新处罚"}',
                "first_seen_at": observed,
                "last_seen_at": observed,
                "seen_count": 1,
            },
        ]
        with tempfile.TemporaryDirectory() as directory:
            path = XlsxExporter.export_history_penalties(
                rows,
                Path(directory) / "penalties.xlsx",
            )
            workbook = load_workbook(path, read_only=False, data_only=True)
            sheet = workbook["行政处罚历史"]
            values = [cell.value for cell in sheet[5]] + [cell.value for cell in sheet[6]]

        self.assertIn("旧处罚", values)
        self.assertIn("新处罚", values)
        self.assertEqual(workbook["导出说明"]["B6"].value[:4], "只增不减")


if __name__ == "__main__":
    unittest.main()
