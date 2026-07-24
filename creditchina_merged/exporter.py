"""XLSX 导出：本次采集、历史行政处罚和全部历史信息。"""

from __future__ import annotations

import json
import os
import tempfile
from collections import OrderedDict
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping, Optional, Sequence

from .crawler import EnterpriseRecord


SECTION_SHEETS = OrderedDict(
    (
        ("基本信息", "基本信息"),
        ("行政许可", "行政许可"),
        ("行政处罚", "行政处罚"),
        ("守信红名单", "守信红名单"),
        ("重点关注名单", "重点关注名单"),
        ("黑名单", "黑名单"),
    )
)


def _load_openpyxl() -> Dict[str, Any]:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError("XLSX 导出需要先安装 openpyxl") from exc
    return {
        "Workbook": Workbook,
        "Alignment": Alignment,
        "Border": Border,
        "Font": Font,
        "PatternFill": PatternFill,
        "Side": Side,
        "get_column_letter": get_column_letter,
    }


def _json_value(value: Any) -> Dict[str, Any]:
    if isinstance(value, dict):
        return dict(value)
    if isinstance(value, (bytes, bytearray)):
        value = value.decode("utf-8", errors="replace")
    if isinstance(value, str):
        try:
            parsed = json.loads(value)
            return dict(parsed) if isinstance(parsed, dict) else {"原始值": parsed}
        except json.JSONDecodeError:
            return {"原始值": value}
    return {"原始值": value}


def _safe_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (datetime, date, int, float, bool)):
        return value
    if isinstance(value, (dict, list, tuple)):
        value = json.dumps(value, ensure_ascii=False, default=str)
    rendered = str(value).replace("\x00", "")
    # 防止数据字段在 Excel 中被当成公式执行。
    if rendered.startswith(("=", "+", "-", "@")):
        return "'" + rendered
    return rendered


class XlsxExporter:
    """生成三类结构化 XLSX 文档。"""

    TITLE_FILL = "155EEF"
    HEADER_FILL = "EAF2FF"
    HEADER_FONT = "1F4FA8"
    LINE_COLOR = "DDE6F4"
    BODY_FONT = "334155"
    MUTED_FONT = "64748B"

    @classmethod
    def _new_workbook(cls) -> Any:
        api = _load_openpyxl()
        workbook = api["Workbook"]()
        workbook.remove(workbook.active)
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        return workbook

    @classmethod
    def _add_overview(
        cls,
        workbook: Any,
        title: str,
        description: str,
        counts: Mapping[str, int],
        company_name: Optional[str] = None,
    ) -> None:
        api = _load_openpyxl()
        sheet = workbook.create_sheet("导出说明", 0)
        sheet.sheet_view.showGridLines = False
        sheet.merge_cells("A1:F2")
        cell = sheet["A1"]
        cell.value = title
        cell.fill = api["PatternFill"]("solid", fgColor=cls.TITLE_FILL)
        cell.font = api["Font"](name="Microsoft YaHei", size=18, bold=True, color="FFFFFF")
        cell.alignment = api["Alignment"](vertical="center", horizontal="left")
        sheet.row_dimensions[1].height = 31
        sheet.row_dimensions[2].height = 18
        sheet["A4"] = "导出范围"
        sheet["B4"] = company_name or "全部历史搜索企业"
        sheet["A5"] = "生成时间"
        sheet["B5"] = datetime.now()
        sheet["B5"].number_format = "yyyy-mm-dd hh:mm:ss"
        sheet["A6"] = "增量规则"
        sheet["B6"] = "只增不减：新内容追加；官网已删除的历史内容继续保留"
        sheet.merge_cells("B6:F6")
        sheet["A8"] = "文档说明"
        sheet["B8"] = description
        sheet.merge_cells("B8:F8")
        for row in (4, 5, 6, 8):
            sheet.cell(row=row, column=1).font = api["Font"](bold=True, color=cls.HEADER_FONT)
            sheet.cell(row=row, column=1).fill = api["PatternFill"]("solid", fgColor=cls.HEADER_FILL)
            sheet.cell(row=row, column=1).alignment = api["Alignment"](vertical="center")
        start = 10
        sheet.cell(start, 1, "数据类别")
        sheet.cell(start, 2, "记录数")
        for index, (label, count) in enumerate(counts.items(), start=start + 1):
            sheet.cell(index, 1, label)
            sheet.cell(index, 2, int(count))
            sheet.cell(index, 2).number_format = "#,##0"
        cls._style_header(sheet, start, 2)
        sheet.column_dimensions["A"].width = 23
        sheet.column_dimensions["B"].width = 34
        for column in ("C", "D", "E", "F"):
            sheet.column_dimensions[column].width = 16
        sheet.freeze_panes = "A4"

    @classmethod
    def _style_header(cls, sheet: Any, row: int, column_count: int) -> None:
        api = _load_openpyxl()
        thin = api["Side"](style="thin", color=cls.LINE_COLOR)
        for column in range(1, column_count + 1):
            cell = sheet.cell(row=row, column=column)
            cell.fill = api["PatternFill"]("solid", fgColor=cls.HEADER_FILL)
            cell.font = api["Font"](name="Microsoft YaHei", size=10, bold=True, color=cls.HEADER_FONT)
            cell.alignment = api["Alignment"](vertical="center", horizontal="left")
            cell.border = api["Border"](bottom=thin)
        sheet.row_dimensions[row].height = 25

    @classmethod
    def _add_data_sheet(
        cls,
        workbook: Any,
        sheet_name: str,
        title: str,
        rows: Sequence[Mapping[str, Any]],
    ) -> None:
        api = _load_openpyxl()
        sheet = workbook.create_sheet(sheet_name[:31])
        sheet.sheet_view.showGridLines = False
        sheet.sheet_properties.tabColor = cls.TITLE_FILL

        columns: List[str] = []
        seen = set()
        for row in rows:
            for key in row:
                rendered = str(key)
                if rendered not in seen:
                    seen.add(rendered)
                    columns.append(rendered)
        if not columns:
            columns = ["说明"]
            rows = [{"说明": "暂无数据"}]

        final_column = api["get_column_letter"](max(1, len(columns)))
        sheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=len(columns))
        title_cell = sheet["A1"]
        title_cell.value = title
        title_cell.fill = api["PatternFill"]("solid", fgColor=cls.TITLE_FILL)
        title_cell.font = api["Font"](name="Microsoft YaHei", size=15, bold=True, color="FFFFFF")
        title_cell.alignment = api["Alignment"](vertical="center", horizontal="left")
        sheet.row_dimensions[1].height = 28
        sheet.row_dimensions[2].height = 16
        header_row = 4
        for column, name in enumerate(columns, start=1):
            sheet.cell(header_row, column, name)
        cls._style_header(sheet, header_row, len(columns))

        thin = api["Side"](style="thin", color="E8EEF7")
        for row_index, row in enumerate(rows, start=header_row + 1):
            for column_index, column_name in enumerate(columns, start=1):
                cell = sheet.cell(row_index, column_index, _safe_cell(row.get(column_name, "")))
                cell.font = api["Font"](name="Microsoft YaHei", size=9, color=cls.BODY_FONT)
                cell.alignment = api["Alignment"](vertical="top", horizontal="left", wrap_text=True)
                cell.border = api["Border"](bottom=thin)
                if isinstance(cell.value, (datetime, date)):
                    cell.number_format = "yyyy-mm-dd hh:mm:ss" if isinstance(cell.value, datetime) else "yyyy-mm-dd"
            sheet.row_dimensions[row_index].height = 24
        sheet.freeze_panes = "A5"
        sheet.auto_filter.ref = f"A4:{final_column}{header_row + len(rows)}"

        for column_index, column_name in enumerate(columns, start=1):
            sample_values = [str(row.get(column_name, "")) for row in rows[:100]]
            longest = max([len(column_name)] + [min(len(value), 80) for value in sample_values])
            width = min(42, max(12, longest * 1.35 + 2))
            if column_name in ("企业名称", "处罚结果", "处罚事由", "原始值"):
                width = max(width, 24)
            sheet.column_dimensions[api["get_column_letter"](column_index)].width = width

    @staticmethod
    def _history_rows(rows: Iterable[Mapping[str, Any]]) -> List[Dict[str, Any]]:
        result = []
        for row in rows:
            payload = _json_value(row.get("raw_json"))
            item: Dict[str, Any] = {
                "企业名称": row.get("company_name", ""),
                "记录键": row.get("record_key", ""),
                "首次发现时间": row.get("first_seen_at", ""),
                "最后发现时间": row.get("last_seen_at", ""),
                "历史出现次数": row.get("seen_count", 1),
            }
            item.update(payload)
            result.append(item)
        return result

    @classmethod
    def _save(cls, workbook: Any, destination: Path) -> Path:
        destination = Path(destination)
        destination.parent.mkdir(parents=True, exist_ok=True)
        file_descriptor, temporary = tempfile.mkstemp(
            prefix=destination.stem + ".",
            suffix=".xlsx",
            dir=str(destination.parent),
        )
        os.close(file_descriptor)
        try:
            workbook.save(temporary)
            os.replace(temporary, destination)
        except Exception:
            try:
                os.unlink(temporary)
            except FileNotFoundError:
                pass
            raise
        return destination

    @classmethod
    def export_current_company(cls, record: EnterpriseRecord, destination: Path) -> Path:
        workbook = cls._new_workbook()
        sections: Mapping[str, Sequence[Mapping[str, Any]]] = OrderedDict(
            (
                ("基本信息", [{"企业名称": record.name, "encryStr": record.encry_str, **record.basic}]),
                ("行政许可", record.permissions),
                ("行政处罚", record.penalties),
                ("守信红名单", record.red_list),
                ("重点关注名单", record.watch_list),
                ("黑名单", record.black_list),
            )
        )
        counts = OrderedDict((section, len(rows)) for section, rows in sections.items())
        cls._add_overview(
            workbook,
            f"{record.name} · 本次采集全部信息",
            "仅包含本次采集返回的企业基本信息和五类信用记录。",
            counts,
            company_name=record.name,
        )
        for section, rows in sections.items():
            cls._add_data_sheet(workbook, SECTION_SHEETS[section], f"{record.name} · {section}", rows)
        if record.errors:
            cls._add_data_sheet(
                workbook,
                "采集错误",
                f"{record.name} · 采集错误",
                [{"栏目": key, "错误信息": value} for key, value in record.errors.items()],
            )
        return cls._save(workbook, destination)

    @classmethod
    def export_history_penalties(
        cls,
        rows: Sequence[Mapping[str, Any]],
        destination: Path,
        company_name: Optional[str] = None,
    ) -> Path:
        workbook = cls._new_workbook()
        history_rows = cls._history_rows(rows)
        cls._add_overview(
            workbook,
            "历史搜索行政处罚",
            "包含当前与所有历史采集中曾经出现的行政处罚，已从官网删除的记录仍然保留。",
            OrderedDict((("行政处罚历史", len(history_rows)),)),
            company_name=company_name,
        )
        cls._add_data_sheet(workbook, "行政处罚历史", "历史搜索 · 行政处罚", history_rows)
        return cls._save(workbook, destination)

    @classmethod
    def export_history_all(
        cls,
        rows: Sequence[Mapping[str, Any]],
        destination: Path,
        company_name: Optional[str] = None,
    ) -> Path:
        grouped: Dict[str, List[Mapping[str, Any]]] = {section: [] for section in SECTION_SHEETS}
        for row in rows:
            section = str(row.get("section_name", ""))
            if section in grouped:
                grouped[section].append(row)

        workbook = cls._new_workbook()
        counts = OrderedDict((section, len(grouped[section])) for section in SECTION_SHEETS)
        cls._add_overview(
            workbook,
            "历史搜索全部信息",
            "包含当前与所有历史采集中曾经出现的六类数据，同一内容保留首次与最后发现时间。",
            counts,
            company_name=company_name,
        )
        for section, sheet_name in SECTION_SHEETS.items():
            cls._add_data_sheet(
                workbook,
                sheet_name,
                f"历史搜索 · {section}",
                cls._history_rows(grouped[section]),
            )
        return cls._save(workbook, destination)
